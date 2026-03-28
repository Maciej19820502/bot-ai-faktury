"""Invoice lookup in Excel file."""

import os
import re
from dataclasses import dataclass, asdict
from typing import Optional

import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "invoices.xlsx")


@dataclass
class InvoiceRecord:
    numer_faktury: str
    dostawca: str
    kwota: float
    waluta: str
    data_wystawienia: str
    termin_platnosci: str
    status: str
    data_platnosci: str

    def to_dict(self) -> dict:
        return asdict(self)

    def to_display(self) -> str:
        lines = [
            f"Numer faktury: {self.numer_faktury}",
            f"Dostawca: {self.dostawca}",
            f"Kwota: {self.kwota} {self.waluta}",
            f"Data wystawienia: {self.data_wystawienia}",
            f"Termin płatności: {self.termin_platnosci}",
            f"Status: {self.status}",
        ]
        if self.data_platnosci:
            lines.append(f"Data płatności: {self.data_platnosci}")
        return "\n".join(lines)


def _normalize_invoice_number(num: str) -> str:
    """Normalize invoice number for comparison.

    FV/2025/001, FV-2025-001, FV 2025 001 all become FV2025001.
    """
    if not num:
        return ""
    return re.sub(r"[\s/\-_.]", "", num).upper().strip()


def search_invoice(invoice_number: str) -> Optional[InvoiceRecord]:
    """Search for an invoice by number (fuzzy match on separators)."""
    if not os.path.exists(EXCEL_PATH):
        return None

    normalized_query = _normalize_invoice_number(invoice_number)
    if not normalized_query:
        return None

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    for row in rows:
        if not row or not row[0]:
            continue
        normalized_row = _normalize_invoice_number(str(row[0]))
        if normalized_row == normalized_query:
            return InvoiceRecord(
                numer_faktury=str(row[0]),
                dostawca=str(row[1] or ""),
                kwota=float(row[2] or 0),
                waluta=str(row[3] or "PLN"),
                data_wystawienia=str(row[4] or ""),
                termin_platnosci=str(row[5] or ""),
                status=str(row[6] or ""),
                data_platnosci=str(row[7] or ""),
            )

    # Try partial match (query contained in invoice number or vice versa)
    for row in rows:
        if not row or not row[0]:
            continue
        normalized_row = _normalize_invoice_number(str(row[0]))
        if normalized_query in normalized_row or normalized_row in normalized_query:
            return InvoiceRecord(
                numer_faktury=str(row[0]),
                dostawca=str(row[1] or ""),
                kwota=float(row[2] or 0),
                waluta=str(row[3] or "PLN"),
                data_wystawienia=str(row[4] or ""),
                termin_platnosci=str(row[5] or ""),
                status=str(row[6] or ""),
                data_platnosci=str(row[7] or ""),
            )

    return None


def get_overdue_invoices_for_client(client_name: str) -> list[InvoiceRecord]:
    """Find all overdue invoices for a given client/supplier."""
    if not os.path.exists(EXCEL_PATH):
        return []

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    results = []
    for row in rows:
        if not row or not row[1]:
            continue
        if str(row[1]).strip().lower() == client_name.strip().lower():
            if str(row[6] or "").strip().lower() == "przeterminowana":
                results.append(
                    InvoiceRecord(
                        numer_faktury=str(row[0]),
                        dostawca=str(row[1] or ""),
                        kwota=float(row[2] or 0),
                        waluta=str(row[3] or "PLN"),
                        data_wystawienia=str(row[4] or ""),
                        termin_platnosci=str(row[5] or ""),
                        status=str(row[6] or ""),
                        data_platnosci=str(row[7] or ""),
                    )
                )
    return results
